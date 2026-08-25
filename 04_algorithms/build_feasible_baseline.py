from __future__ import annotations

import csv
import math
import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl


HOURS = 2407
TOL = 1e-8


def read_rows(path: Path, sheet: str | None = None) -> list[dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0] if sheet is None else sheet]
    rows = worksheet.iter_rows(values_only=True)
    header = list(next(rows))
    return [dict(zip(header, row)) for row in rows]


def overlaps(start: int, duration_hours: float):
    end = start + duration_hours
    for hour in range(start, min(2406, math.ceil(end) - 1) + 1):
        overlap = max(0.0, min(end, hour + 1.0) - max(float(start), float(hour)))
        if overlap > 0:
            yield hour, overlap


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the feasible no-migration x_base task schedule.")
    parser.add_argument("--data-dir", required=True, type=Path, help="Directory containing the six input xlsx files.")
    parser.add_argument("--output", type=Path, default=Path("x_base_task_schedule.csv"), help="Output CSV path.")
    args = parser.parse_args()
    raw = args.data_dir
    output_path = args.output
    tasks = read_rows(raw / "workload_trace.xlsx")
    regions = read_rows(raw / "GPU_information.xlsx")
    powers = read_rows(raw / "power_mapping.xlsx")
    region_time = read_rows(raw / "region_time_data.xlsx")
    latency_rows = read_rows(raw / "network_latency.xlsx")

    capacity = {row["Region"]: row for row in regions}
    power = {
        row["TaskType"]: float(row["GPU_Power_MW_per_EquivalentGPU"])
        for row in powers
    }
    non_ai = {
        (row["Region"], int(row["Hour"])): float(row["NonAI_IT_Load_MW"] or 0)
        for row in region_time
    }

    latency = {}
    for row in latency_rows:
        source = row.get("SourceRegion") or row.get("FromRegion") or row.get("Region")
        target = row.get("TargetRegion") or row.get("ToRegion")
        value = row.get("Latency_ms") or row.get("NetworkLatency_ms")
        if source is not None and target is not None and value is not None:
            latency[(source, target)] = float(value)

    gpu_load = {region: [0.0] * HOURS for region in capacity}
    ai_it_load = {region: [0.0] * HOURS for region in capacity}
    schedule: dict[str, dict] = {}
    unscheduled: list[str] = []

    def feasible(task: dict, start: int) -> bool:
        region = task["SourceRegion"]
        demand = float(task["GPU_Demand"])
        unit_power = power[task["TaskType"]]
        duration = float(task["EstimatedDuration_min"]) / 60.0
        if start < float(task["EarliestStartHour"]) - TOL:
            return False
        if start + duration > min(float(task["LatestFinishHour"]), 2406.0) + TOL:
            return False
        if latency.get((region, region), math.inf) > float(task["MaxLatency_ms"]) + TOL:
            return False
        for hour, overlap in overlaps(start, duration):
            gpu = gpu_load[region][hour] + demand * overlap
            ai_it = ai_it_load[region][hour] + demand * unit_power * overlap
            it = non_ai[(region, hour)] + ai_it
            facility = it * float(capacity[region]["PUE"])
            if gpu > float(capacity[region]["Available_GPU"]) + TOL:
                return False
            if it > float(capacity[region]["Max_IT_Power_MW"]) + TOL:
                return False
            if facility > float(capacity[region]["Max_Facility_Power_MW"]) + TOL:
                return False
        return True

    def place(task: dict, start: int) -> None:
        region = task["SourceRegion"]
        demand = float(task["GPU_Demand"])
        unit_power = power[task["TaskType"]]
        duration = float(task["EstimatedDuration_min"]) / 60.0
        for hour, overlap in overlaps(start, duration):
            gpu_load[region][hour] += demand * overlap
            ai_it_load[region][hour] += demand * unit_power * overlap
        local_latency = latency.get((region, region), 0.0)
        schedule[str(task["TaskID"])] = {
            "TaskID": task["TaskID"],
            "TaskType": task["TaskType"],
            "SourceRegion": region,
            "TargetRegion": region,
            "ArrivalHour": int(task["ArrivalHour"]),
            "StartHour": start,
            "EndHour": start + duration,
            "LatestFinishHour": float(task["LatestFinishHour"]),
            "GPU_Demand": demand,
            "NetworkLatency_ms": local_latency,
        }

    real_time = [task for task in tasks if task["TaskType"] == "RealTimeInference"]
    flexible = [task for task in tasks if task["TaskType"] != "RealTimeInference"]

    for task in sorted(real_time, key=lambda x: (int(x["ArrivalHour"]), str(x["TaskID"]))):
        start = int(task["ArrivalHour"])
        if not feasible(task, start):
            unscheduled.append(str(task["TaskID"]))
        else:
            place(task, start)

    flexible.sort(
        key=lambda x: (
            float(x["LatestFinishHour"]),
            int(x["ArrivalHour"]),
            -float(x["GPU_Demand"]) * float(x["EstimatedDuration_min"]),
            str(x["TaskID"]),
        )
    )
    for task in flexible:
        duration = float(task["EstimatedDuration_min"]) / 60.0
        earliest = max(int(task["ArrivalHour"]), int(task["EarliestStartHour"]))
        latest_start = math.floor(min(float(task["LatestFinishHour"]), 2406.0) - duration + TOL)
        chosen = None
        for start in range(earliest, latest_start + 1):
            if feasible(task, start):
                chosen = start
                break
        if chosen is None:
            unscheduled.append(str(task["TaskID"]))
        else:
            place(task, chosen)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "TaskID", "TaskType", "SourceRegion", "TargetRegion", "ArrivalHour",
        "StartHour", "EndHour", "LatestFinishHour", "GPU_Demand", "NetworkLatency_ms",
    ]
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for task_id in sorted(schedule, key=lambda x: int(x) if x.isdigit() else x):
            writer.writerow(schedule[task_id])

    max_gpu = (0.0, "", -1, 0.0, 0.0)
    max_it = (0.0, "", -1, 0.0, 0.0)
    max_facility = (0.0, "", -1, 0.0, 0.0)
    violations = 0
    for region, row in capacity.items():
        for hour in range(HOURS):
            gpu = gpu_load[region][hour]
            it = non_ai[(region, hour)] + ai_it_load[region][hour]
            facility = it * float(row["PUE"])
            gpu_ratio = gpu / float(row["Available_GPU"])
            it_ratio = it / float(row["Max_IT_Power_MW"])
            facility_ratio = facility / float(row["Max_Facility_Power_MW"])
            max_gpu = max(max_gpu, (gpu_ratio, region, hour, gpu, float(row["Available_GPU"])))
            max_it = max(max_it, (it_ratio, region, hour, it, float(row["Max_IT_Power_MW"])))
            max_facility = max(
                max_facility,
                (facility_ratio, region, hour, facility, float(row["Max_Facility_Power_MW"])),
            )
            violations += int(gpu_ratio > 1 + TOL or it_ratio > 1 + TOL or facility_ratio > 1 + TOL)

    delayed = sum(row["StartHour"] > row["ArrivalHour"] for row in schedule.values())
    print(f"scheduled={len(schedule)} unscheduled={len(unscheduled)} delayed={delayed}")
    print(f"capacity_violations={violations}")
    print(f"max_gpu={max_gpu}")
    print(f"max_it={max_it}")
    print(f"max_facility={max_facility}")
    if unscheduled:
        print("unscheduled_ids=" + ",".join(unscheduled[:20]))
        raise SystemExit(2)
    if violations:
        raise SystemExit(3)


if __name__ == "__main__":
    main()
